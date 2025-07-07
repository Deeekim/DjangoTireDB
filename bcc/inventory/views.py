import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Max, Sum, F, Case, When, IntegerField, Q
from django.utils.timezone import now
from .forms import *
from .models import *
from .dependency import *
from collections import defaultdict

from django.http import JsonResponse, HttpResponse
from django.core.cache import cache
from django.views.decorators.csrf import csrf_exempt

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.urls import reverse

from datetime import date

def home(request):
    return render(request, 'inventory/base_template.html')

# Each section is segmented per purpose. These are separated using three lines of hashtags
# The first section is designated to view functions for Tire related transactions
# The second section is designated to view functions for Magwheel related transactions
# The third section is designated to view functions related to the Masterlist, and its updating from Google Sheets
# The fourth section is currently assigned for miscellaneous functions

########################################################################################################################
########################################################################################################################
########################################################################################################################


def viewAllTires(request):
    tireTransactions = Tire.objects.all()
    context = { 'tireTransactions': tireTransactions[::-1] }
    return render(request, 'inventory/tireTransaction_viewAll.html', context)

def viewSummaryTires(request):
    tires = Tire.objects.all()

    # Group and calculate total inventory and last transaction date
    grouped_inventory = (
        tires
        .annotate(
            inventory_effect=Case(
                When(transaction_type='IN', then=F('qty')),
                When(transaction_type='OUT', then=-F('qty')),
                default=0,
                output_field=IntegerField()
            )
        )
        .values('location', 'brand', 'size', 'model', 'specs', 'supplier', 'load', 'remarks')
        .annotate(
            total_inventory=Sum('inventory_effect'),
            last_transaction_date=Max('date')  # Get the most recent date for each group
        )
        .order_by('brand', 'size', 'model', 'specs')  # Optional for consistent grouping order
    )

    context = {
        'grouped_inventory': grouped_inventory,
    }

    return render(request, 'inventory/tireTransaction_viewSummary.html', context)
 
def queryTireTransaction(request):
    # Start with all transactions
    tires = Tire.objects.all()

    # Define default filter values
    filters = {
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'location': request.GET.get('location', ''),
        'transaction_type': request.GET.get('transaction_type', ''),
        'brand': request.GET.get('brand', ''),
        'size': request.GET.get('size', ''),
        'model': request.GET.get('model', ''),
        'specs': request.GET.get('specs', ''),
        'supplier': request.GET.get('supplier', ''),
        'notes': request.GET.get('notes',''),
    }

    # Filter transactions by date range
    if filters['start_date'] and filters['end_date']:
        tires = tires.filter(date__range=[filters['start_date'], filters['end_date']])
    elif filters['start_date']:
        tires = tires.filter(date__gte=filters['start_date'])
    elif filters['end_date']:
        tires = tires.filter(date__lte=filters['end_date'])

    # Apply filters if queries are provided
    if filters['location']:
        tires = tires.filter(location__icontains=filters['location'])
    if filters['transaction_type']:
        tires = tires.filter(transaction_type__icontains=filters['transaction_type'])
    if filters['brand']:
        tires = tires.filter(brand__icontains=filters['brand'])
    if filters['size']:
        tires = tires.filter(size__icontains=filters['size'])
    if filters['model']:
        tires = tires.filter(model__icontains=filters['model'])
    if filters['specs']:
        tires = tires.filter(specs__icontains=filters['specs'])
    if filters['supplier']:
        tires = tires.filter(supplier__icontains=filters['supplier'])
    
    # Sort by date (earliest first)
    tires = tires.order_by('date')  # Change here: order by date from earliest to latest

    # Initialize the cumulative quantity dictionary
    cumulative_qty = {}
    tires_list = []

    for tire in tires:
        # Create a key for grouping by Brand, Size, Model, Specs
        key = (tire.location, tire.brand, tire.size, tire.model, tire.specs, tire.load, tire.supplier)

        # Add or subtract quantity based on the transaction type
        if tire.transaction_type == "IN":
            qty_change = tire.qty
        elif tire.transaction_type == "OUT":
            qty_change = -tire.qty
        else:
            qty_change = 0  # In case of other transaction types, you can handle this accordingly

        # Initialize cumulative quantity for this group if not already done
        if key not in cumulative_qty:
            cumulative_qty[key] = 0

        # Update cumulative quantity
        cumulative_qty[key] += qty_change

        # Add the transaction data along with the cumulative inventory quantity
        tires_list.append({
            'tire': tire,
            'inventory_qty': cumulative_qty[key]
        })

    # Pass transactions and current filters to the template
    context = {
        'tires': tires_list[::-1],  # Pass the modified transaction data
        'filters': filters,
    }

    return render(request, 'inventory/tireTransaction_queryTransaction.html', context)

def createTireTransaction(request):
    # Determine if the request is part of a Full Transaction workflow
    is_full_transaction = request.GET.get('full_transaction',  'false').lower() == 'true'
    # Create vs Edit Transaction
    isediting = False

    # Choices creation
    TIRES_BRAND_CHOICES = get_tires_brand_choices()
    TIRES_BRAND_SIZE_CHOICES = get_tires_brand_size_choices()
    TIRES_MODEL_SPECS_CHOICES = get_tires_model_specs_choices()
    TIRES_SPECS_LOAD_CHOICES = get_tires_specs_load_choices()
    TIRES_SPECS_REMARKS_CHOICES = get_tires_specs_remarks_choices()

    if request.method == "POST":
        form = TireTransactionForm(request.POST)
        if form.is_valid():
            tire_data = form.cleaned_data
            for field, value in tire_data.items():
                if field == 'invoice_number' and value is None:
                    tire_data[field] = 0
                elif field != 'invoice_number' and value is None:
                    tire_data[field] = ""


            if is_full_transaction:
                # Add to session for Full Transaction
                transaction_items = request.session.get('transaction_items', [])
                tire_data_serializable = {
                    key: (value.isoformat() if isinstance(value, date) else value)
                    for key, value in tire_data.items()
                }
                transaction_items.append({
                    'type': 'tire',
                    'data': tire_data_serializable,
                    'display': f"{tire_data['brand']}  {tire_data['model']}  {tire_data['specs']}  {tire_data['load']} - {tire_data['supplier']}",
                })
                
                # Update session with modified transaction_items
                request.session['transaction_items'] = transaction_items
                request.session.modified = True  # Ensure session changes are saved
                
                return redirect('inventory:full-transaction-view')
            
            else:
                # Save immediately for single transaction
                Tire.objects.create(**form.cleaned_data)
                return redirect('inventory:all-tire-transaction')
    else:
        form = TireTransactionForm()

    context = {
        'form': form,
        'isediting': isediting,
        'is_full_transaction': is_full_transaction,
        'brand_choices': TIRES_BRAND_CHOICES,
        'brand_size_choices': json.dumps(TIRES_BRAND_SIZE_CHOICES),
        'model_specs_choices': json.dumps(TIRES_MODEL_SPECS_CHOICES),
        'specs_load_choices': json.dumps(TIRES_SPECS_LOAD_CHOICES),
        'specs_remarks_choices': json.dumps(TIRES_SPECS_REMARKS_CHOICES),
    }
    return render(request, 'inventory/tireTransaction_form.html', context)
    
def editTireTransaction(request, pk):
    # Get transaction for editing
    tireTransaction = Tire.objects.get(id=pk)

    # Create vs Edit Transaction
    isediting = True

    # Choices creation
    TIRES_BRAND_CHOICES = get_tires_brand_choices()
    TIRES_BRAND_SIZE_CHOICES = get_tires_brand_size_choices()
    TIRES_MODEL_SPECS_CHOICES = get_tires_model_specs_choices()
    TIRES_SPECS_LOAD_CHOICES = get_tires_specs_load_choices()
    TIRES_SPECS_REMARKS_CHOICES = get_tires_specs_remarks_choices()

    if request.method == "POST":
        form = TireTransactionForm(request.POST, instance=tireTransaction)
        if form.is_valid():
            form.save()
            return redirect('inventory:all-tire-transaction')
    else:
        form = TireTransactionForm(instance=tireTransaction)

    context = {
        'form': form,
        'isediting': isediting,
        'tireTransaction': tireTransaction,
        'brand_choices': TIRES_BRAND_CHOICES,
        'brand_size_choices': json.dumps(TIRES_BRAND_SIZE_CHOICES),
        'model_specs_choices': json.dumps(TIRES_MODEL_SPECS_CHOICES),
        'specs_load_choices': json.dumps(TIRES_SPECS_LOAD_CHOICES),
        'specs_remarks_choices': json.dumps(TIRES_SPECS_REMARKS_CHOICES),
    }
    return render(request, 'inventory/tireTransaction_form.html', context)

def deleteTireTransaction(request, pk):
    tireTransaction = Tire.objects.get(id=pk)
    if request.method == "GET":
        context = { 'tireTransaction': tireTransaction }
        return render(request, "inventory/tireTransaction_delete.html", context)
    elif request.method == "POST":
        tireTransaction.delete()
        return redirect('inventory:all-tire-transaction')


########################################################################################################################
########################################################################################################################
########################################################################################################################


def viewAllMagwheels(request):
    magwheelTransactions = Magwheel.objects.all()
    context = { 'magwheelTransactions': magwheelTransactions[::-1] }
    return render(request, 'inventory/magwheelTransaction_viewAll.html', context)

def viewSummaryMagwheels(request):
    magwheels = Magwheel.objects.all()

    # Group and calculate total inventory and last transaction date
    grouped_inventory = (
        magwheels
        .annotate(
            inventory_effect=Case(
                When(transaction_type='IN', then=F('qty')),
                When(transaction_type='OUT', then=-F('qty')),
                default=0,
                output_field=IntegerField()
            )
        )
        .values('location', 'brand', 'size', 'model', 'pcd1', 'pcd2', 'offset', 'bore', 'color', 'supplier')
        .annotate(
            total_inventory=Sum('inventory_effect'),
            last_transaction_date=Max('date')  # Get the most recent date for each group
        )
        .order_by('brand', 'size', 'model', 'pcd1', 'offset', 'color')  # Optional for consistent grouping order
    )

    context = {
        'grouped_inventory': grouped_inventory,
    }

    return render(request, 'inventory/magwheelTransaction_viewSummary.html', context)

def queryMagwheelTransaction(request):
    # Start with all transactions
    magwheels = Magwheel.objects.all()

    # Define default filter values
    filters = {
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'location': request.GET.get('location', ''),
        'transaction_type': request.GET.get('transaction_type', ''),
        'brand': request.GET.get('brand', ''),
        'size': request.GET.get('size', ''),
        'model': request.GET.get('model', ''),
        'pcd1': request.GET.get('pcd1', ''),
        'pcd2': request.GET.get('pcd2', ''),
        'offset': request.GET.get('offset', ''),
        'color': request.GET.get('color',''),
        'supplier': request.GET.get('supplier', ''),
        'notes': request.GET.get('notes',''),
    }

    # Filter transactions by date range
    if filters['start_date'] and filters['end_date']:
        magwheels = magwheels.filter(date__range=[filters['start_date'], filters['end_date']])
    elif filters['start_date']:
        magwheels = magwheels.filter(date__gte=filters['start_date'])
    elif filters['end_date']:
        magwheels = magwheels.filter(date__lte=filters['end_date'])

    # Apply filters if queries are provided
    if filters['location']:
        magwheels = magwheels.filter(location__icontains=filters['location'])
    if filters['transaction_type']:
        magwheels = magwheels.filter(transaction_type__icontains=filters['transaction_type'])
    if filters['brand']:
        magwheels = magwheels.filter(brand__icontains=filters['brand'])
    if filters['size']:
        magwheels = magwheels.filter(size__icontains=filters['size'])
    if filters['model']:
        magwheels = magwheels.filter(model__icontains=filters['model'])
    if filters['pcd1']:
        magwheels = magwheels.filter(pcd1__icontains=filters['pcd1'])
    if filters['pcd2']:
        magwheels = magwheels.filter(pcd2__icontains=filters['pcd2'])
    if filters['offset']:
        magwheels = magwheels.filter(offset__icontains=filters['offset'])
    if filters['color']:
        magwheels = magwheels.filter(color__icontains=filters['color'])
    if filters['supplier']:
        magwheels = magwheels.filter(supplier__icontains=filters['supplier'])
    
    # Sort by date (earliest first)
    magwheels = magwheels.order_by('date')  # Change here: order by date from earliest to latest

    # Initialize the cumulative quantity dictionary
    cumulative_qty = {}
    magwheels_list = []

    for magwheel in magwheels:
        # Create a key for grouping by Brand, Size, Model, Specs
        key = (magwheel.location, magwheel.brand, magwheel.size, magwheel.model, magwheel.pcd1, magwheel.pcd2, magwheel.offset, magwheel.bore, magwheel.color)

        # Add or subtract quantity based on the transaction type
        if magwheel.transaction_type == "IN":
            qty_change = magwheel.qty
        elif magwheel.transaction_type == "OUT":
            qty_change = -magwheel.qty
        else:
            qty_change = 0  # In case of other transaction types, you can handle this accordingly

        # Initialize cumulative quantity for this group if not already done
        if key not in cumulative_qty:
            cumulative_qty[key] = 0

        # Update cumulative quantity
        cumulative_qty[key] += qty_change

        # Add the transaction data along with the cumulative inventory quantity
        magwheels_list.append({
            'magwheel': magwheel,
            'inventory_qty': cumulative_qty[key]
        })

    # Pass transactions and current filters to the template
    context = {
        'magwheels': magwheels_list[::-1],  # Pass the modified transaction data
        'filters': filters,
    }

    return render(request, 'inventory/magwheelTransaction_queryTransaction.html', context)

def createMagwheelTransaction(request):
    # Determine if the request is part of a Full Transaction workflow
    is_full_transaction = request.GET.get('full_transaction',  'false').lower() == 'true'
    
    # Create vs Edit Transaction
    isediting = False

    # Choices creation
    MAGWHEELS_BRAND_CHOICES = get_magwheels_brand_choices()
    MAGWHEELS_BRAND_SIZE_CHOICES = get_magwheels_brand_size_choices()
    MAGWHEELS_MODEL_PCD1_CHOICES = get_magwheels_model_pcd1_choices()
    MAGWHEELS_MODEL_PCD2_CHOICES = get_magwheels_model_pcd2_choices()
    MAGWHEELS_MODEL_OFFSET_CHOICES = get_magwheels_model_offset_choices()
    MAGWHEELS_MODEL_BORE_CHOICES = get_magwheels_model_bore_choices()
    MAGWHEELS_MODEL_COLOR_CHOICES = get_magwheels_model_color_choices()

    if request.method == "POST":
        form = MagwheelTransactionForm(request.POST)
        if form.is_valid():
            magwheel_data = form.cleaned_data

            if is_full_transaction:
                # Add to session for Full Transaction
                transaction_items = request.session.get('transaction_items', [])
                magwheel_data_serializable = {
                    key: (value.isoformat() if isinstance(value, date) else value)
                    for key, value in magwheel_data.items()
                }
                transaction_items.append({
                    'type': 'magwheel',
                    'data': magwheel_data_serializable,
                    'display': f"{magwheel_data['brand']}  {magwheel_data['model']}  {magwheel_data['size']}  {magwheel_data['pcd1']}/{magwheel_data['pcd2']}  {magwheel_data['offset']}  {magwheel_data['bore']} - {magwheel_data['supplier']}",
                })

                # Update session with modified transaction_items
                request.session['transaction_items'] = transaction_items
                request.session.modified = True # Ensure session changes are saved

                return redirect('inventory:full-transaction-view')
            
            else:
                # Save immediately for single transaction
                Magwheel.objects.create(**form.cleaned_data)
                return redirect('inventory:all-magwheel-transaction')
            
    else:
        form = MagwheelTransactionForm()

    context = {
        'form': form,
        'isediting': isediting,
        'is_full_transaction': is_full_transaction,
        'brand_choices': MAGWHEELS_BRAND_CHOICES,
        'brand_size_choices': json.dumps(MAGWHEELS_BRAND_SIZE_CHOICES),
        'model_pcd1_choices': json.dumps(MAGWHEELS_MODEL_PCD1_CHOICES),
        'model_pcd2_choices': json.dumps(MAGWHEELS_MODEL_PCD2_CHOICES),
        'model_offset_choices': json.dumps(MAGWHEELS_MODEL_OFFSET_CHOICES),
        'model_bore_choices': json.dumps(MAGWHEELS_MODEL_BORE_CHOICES),
        'model_color_choices': json.dumps(MAGWHEELS_MODEL_COLOR_CHOICES),
    }
    return render(request, 'inventory/magwheelTransaction_form.html', context)

def editMagwheelTransaction(request, pk):
    # Get transaction for editing
    magwheelTransaction = Magwheel.objects.get(id=pk)
    
    # Create vs Edit Transaction
    isediting = True

    # Choices creation
    MAGWHEELS_BRAND_CHOICES = get_magwheels_brand_choices()
    MAGWHEELS_BRAND_SIZE_CHOICES = get_magwheels_brand_size_choices()
    MAGWHEELS_MODEL_PCD1_CHOICES = get_magwheels_model_pcd1_choices()
    MAGWHEELS_MODEL_PCD2_CHOICES = get_magwheels_model_pcd2_choices()
    MAGWHEELS_MODEL_OFFSET_CHOICES = get_magwheels_model_offset_choices()
    MAGWHEELS_MODEL_BORE_CHOICES = get_magwheels_model_bore_choices()
    MAGWHEELS_MODEL_COLOR_CHOICES = get_magwheels_model_color_choices()

    if request.method == "POST":
        form = MagwheelTransactionForm(request.POST, instance=magwheelTransaction)
        if form.is_valid():
            form.save()
            return redirect('inventory:all-magwheel-transaction')
    else:
        form = MagwheelTransactionForm(instance=magwheelTransaction)

    context = {
        'form': form,
        'isediting': isediting,
        'magwheelTransaction': magwheelTransaction,
        'brand_choices': MAGWHEELS_BRAND_CHOICES,
        'brand_size_choices': json.dumps(MAGWHEELS_BRAND_SIZE_CHOICES),
        'model_pcd1_choices': json.dumps(MAGWHEELS_MODEL_PCD1_CHOICES),
        'model_pcd2_choices': json.dumps(MAGWHEELS_MODEL_PCD2_CHOICES),
        'model_offset_choices': json.dumps(MAGWHEELS_MODEL_OFFSET_CHOICES),
        'model_bore_choices': json.dumps(MAGWHEELS_MODEL_BORE_CHOICES),
        'model_color_choices': json.dumps(MAGWHEELS_MODEL_COLOR_CHOICES),
    }
    return render(request, 'inventory/magwheelTransaction_form.html', context)

def deleteMagwheelTransaction(request, pk):
    magwheelTransaction = Magwheel.objects.get(id=pk)
    if request.method == "GET":
        context = { 'magwheelTransaction': magwheelTransaction }
        return render(request, "inventory/magwheelTransaction_delete.html", context)
    elif request.method == "POST":
        magwheelTransaction.delete()
        return redirect('inventory:all-magwheel-transaction')


########################################################################################################################
########################################################################################################################
########################################################################################################################


def viewMasterlist(request):
    # Get all rows in masterlist
    masterlist = get_masterlist_data()
    
    # Ensure column names are sanitized before converting to dictionaries
    masterlist.columns = [col.replace(' / ', '_').replace(' ', '_') for col in masterlist.columns]

    # Convert the DataFrame to a list of dictionaries
    masterlist_dict = masterlist.to_dict('records')

    # Pass the dictionary to the context
    context = {
        'masterlist': masterlist_dict
    }

    # Render the template
    return render(request, 'inventory/masterlist.html', context)

@csrf_exempt
def refresh_tires(request):
    if request.method == "POST":  # Ensure it's a POST request
        # Delete the cached data
        for key in ["masterlist_data", "tires_masterlist", "tires_brand_choices", "tires_brand_size_choices", "tires_model_color_choices", "tires_model_pcd1_choices", "tires_model_pcd2_choices", "tires_model_offset_choices", "tires_model_bore_choices", "tires_model_color_choices"]:
            cache.delete(key)

        messages.success(request, "Cache successfully invalidated.")

        return JsonResponse({"status": "Cache invalidated"})
    return JsonResponse({"status": "Invalid method"}, status=405)

@csrf_exempt
def refresh_magwheels(request):
    if request.method == "POST":  # Ensure it's a POST request
        # Delete the cached data
        for key in ["masterlist_data", "magwheels_masterlist", "magwheels_brand_choices", "magwheels_brand_size_choices", "magwheels_model_color_choices", "magwheels_model_pcd1_choices", "magwheels_model_pcd2_choices", "magwheels_model_offset_choices", "magwheels_model_bore_choices", "magwheels_model_color_choices"]:
            cache.delete(key)
        
        messages.success(request, "Cache successfully invalidated.")

        return JsonResponse({"status": "Cache invalidated"})
    return JsonResponse({"status": "Invalid method"}, status=405)


########################################################################################################################
########################################################################################################################
########################################################################################################################


def fullTransactionView(request):
    # Get transaction items from the session
    transaction_items = request.session.get('transaction_items', [])

    if request.method == "POST":
        if "add_tire" in request.POST:
            return redirect(f"{reverse('inventory:create-tire-transaction')}?full_transaction=True")
        elif "add_magwheel" in request.POST:
            return redirect(f"{reverse('inventory:create-magwheel-transaction')}?full_transaction=True")
        elif "finalize" in request.POST:
            # Process the items and save to the database
            for item in transaction_items:
                if item['type'] == 'tire':
                    Tire.objects.create(**item['data'])
                elif item['type'] == 'magwheel':
                    Magwheel.objects.create(**item['data'])
            # Clear session after saving
            request.session.pop('transaction_items', None)
            return redirect('inventory:summary-tire-transaction')
        elif "clear_transaction" in request.POST:
            # Clear the transaction items from the session
            request.session.pop('transaction_items', None)
            request.session.modified = True  # Ensure session changes are saved
            return redirect('inventory:full-transaction-view')
        elif "delete_item" in request.POST:
            # Remove a specific item by its index
            delete_index = int(request.POST.get('delete_index', -1))
            if 0 <= delete_index < len(transaction_items):
                del transaction_items[delete_index]  # Remove the item from the list
                request.session['transaction_items'] = transaction_items  # Update the session
                request.session.modified = True  # Mark session as modified to save changes
            return redirect('inventory:full-transaction-view')

    context = {
        'transaction_items': transaction_items,
    }
    return render(request, 'inventory/fullTransaction_form.html', context)

########################################################################################################################
########################################################################################################################
########################################################################################################################

# Version 2.0 below. 
# Transactions, Tires/Magwheels/Services/Accessories Session, Tires/Magwheels/Services/Accessories Save

def v2_create_transaction(request):
    draft_data = request.session.get('transaction_draft')

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            payments = request.session.get('unsaved_payments', [])
            total_payment = sum(Decimal(str(p.get('amount', 0)))for p in payments)

            if total_payment == 0:
                messages.warning(
                    request,
                    "You must add at least one payment. The transaction cannot be saved with ₱0 total."
                )
            
                form = TransactionForm(initial=draft_data)
            else:
                transaction = form.save(commit = False)
                transaction.amount = total_payment
                transaction.save()

                for payment_data in payments:
                    method = payment_data.get('method')

                    # Clean up fields based on method
                    if method != 'Credit Card':
                        payment_data['credit_card_type'] = None
                    if method != 'Debit Card':
                        payment_data['debit_card_bank'] = None
                    if method != 'Cheque':
                        payment_data['cheque_on_date'] = None

                    TransactionPayment.objects.create(transaction=transaction, **payment_data)


                # 3. Save tires from session
                tires = request.session.get('unsaved_tires', [])
                for tire_data in tires:
                    Tire_test.objects.create(transaction=transaction, **tire_data)
                
                # 4. Save magwheels from session
                magwheels = request.session.get('unsaved_magwheels', [])
                for mag_data in magwheels:
                    Magwheel_test.objects.create(transaction=transaction, **mag_data)
                    
                # 5. Clear session
                request.session['unsaved_payments'] = []
                request.session['unsaved_tires'] = []
                request.session['unsaved_magwheels'] = []
                request.session.pop('transaction_draft', None)

                # 6. Redirect
                return redirect('inventory:v2-view-all-transactions')
        else:
            request.session['transaction_draft'] = request.POST.dict()
        
    else:
        form = TransactionForm(initial=draft_data) #  or {'date': now().date()}

    # Ensure session lists exist
    for key in ['unsaved_payments', 'unsaved_tires', 'unsaved_magwheels']:
        if key not in request.session:
            request.session[key] = []
    
    # Print-style preview of what's been added
    added_payments = request.session.get('unsaved_payments',[])
    added_tires = request.session.get('unsaved_tires',[])
    added_mags = request.session.get('unsaved_magwheels',[])
    
    context = {
        'form': form,
        'added_payments': added_payments,
        'added_tires': added_tires,
        'added_magwheels': added_mags,
    }

    return render(request, 'inventory/v2_addTransaction.html', context)

@csrf_exempt
def v2_save_transaction_draft(request):
    if request.method == 'POST':
        request.session['transaction_draft'] = request.POST.dict()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'invalid method'}, status=405)

def v2_cancel_transaction(request):
    request.session.pop('transaction_draft', None)
    request.session.pop('unsaved_tires', None)
    request.session.pop('unsaved_magwheels', None)
    request.session.pop('unsaved_payments', None)
    return redirect('inventory:v2-view-all-transactions')

def v2_viewAllTransactions(request):
    transactions = Transaction.objects.all().order_by('-date')
    return render(request, 'inventory/v2_allTransactions.html', {'transactions': transactions})

def v2_expandTransaction(request, pk):
    transaction = get_object_or_404(Transaction, pk = pk)
    payments = TransactionPayment.objects.filter(transaction = transaction)
    tires = Tire_test.objects.filter(transaction = transaction)
    magwheels = Magwheel_test.objects.filter(transaction = transaction)

    context = {
        'transaction': transaction,
        'payments': payments,
        'tires': tires,
        'magwheels': magwheels,
    }

    return render(request, 'inventory/v2_expandTransaction.html', context)

def v2_editTransaction(request, pk):
    transaction = get_object_or_404(Transaction, pk = pk)
    if request.method == "POST":
        form = TransactionForm(request.POST, instance = transaction)
        if form.is_valid():
            form.save()
            return redirect('inventory:v2-expand-transaction', pk = pk)
        print("Form errors:", form.errors)
    else:
        form = TransactionForm(instance = transaction)

    

    context = {
        'form': form,
        'transaction': transaction,
    }

    return render(request, 'inventory/v2_editTransaction.html', context)

def v2_deleteTransaction(request, pk):
    transaction = get_object_or_404(Transaction, pk = pk)
    transaction.delete()
    return redirect('inventory:v2-view-all-transactions')

def v2_addTire(request, index = None):
    tires = request.session.get('unsaved_tires', [])

    TIRES_BRAND_CHOICES = get_tires_brand_choices()
    TIRES_BRAND_SIZE_CHOICES = get_tires_brand_size_choices()
    TIRES_MODEL_SPECS_CHOICES = get_tires_model_specs_choices()
    TIRES_SPECS_LOAD_CHOICES = get_tires_specs_load_choices()
    TIRES_SPECS_REMARKS_CHOICES = get_tires_specs_remarks_choices()

    if request.method == 'POST':
        form = TireForm(request.POST)
        if form.is_valid():
            tire_data = form.cleaned_data

            if index is not None:
                tires[int(index)] = tire_data
            else:
                tires.append(tire_data)
            
            request.session['unsaved_tires'] = tires
            request.session.modified = True
            return redirect('inventory:v2-create-transaction')

    else:
        if index is not None and 0 <= int(index) < len(tires):
            form = TireForm(initial = tires[int(index)])
            
        else:
            form = TireForm()

    context = {
        'form': form,
        'edit_mode': index is not None,
        'tireTransaction': tires[int(index)] if index is not None and 0 <= int(index) < len(tires) else {},
        'brand_choices': TIRES_BRAND_CHOICES,
        'brand_size_choices': json.dumps(TIRES_BRAND_SIZE_CHOICES),
        'model_specs_choices': json.dumps(TIRES_MODEL_SPECS_CHOICES),
        'specs_load_choices': json.dumps(TIRES_SPECS_LOAD_CHOICES),
        'specs_remarks_choices': json.dumps(TIRES_SPECS_REMARKS_CHOICES),
    }

    return render(request, 'inventory/v2_addTire.html', context)

def v2_deleteTire(request, index):
    tires = request.session.get('unsaved_tires', [])
    if 0 <= int(index) < len(tires):
        del tires[int(index)]
        request.session['unsaved_tires'] = tires
        request.session.modified = True
    return redirect('inventory:v2-create-transaction')

def v2_addMagwheel(request, index = None):
    magwheels = request.session.get('unsaved_magwheels', [])

    # Choices creation
    MAGWHEELS_BRAND_CHOICES = get_magwheels_brand_choices()
    MAGWHEELS_BRAND_SIZE_CHOICES = get_magwheels_brand_size_choices()
    MAGWHEELS_MODEL_PCD1_CHOICES = get_magwheels_model_pcd1_choices()
    MAGWHEELS_MODEL_PCD2_CHOICES = get_magwheels_model_pcd2_choices()
    MAGWHEELS_MODEL_OFFSET_CHOICES = get_magwheels_model_offset_choices()
    MAGWHEELS_MODEL_BORE_CHOICES = get_magwheels_model_bore_choices()
    MAGWHEELS_MODEL_COLOR_CHOICES = get_magwheels_model_color_choices()

    if request.method == 'POST':
        form = MagwheelForm(request.POST)
        if form.is_valid():
            magwheel_data = form.cleaned_data

            if index is not None:
                magwheels[int(index)] = magwheel_data
            else:
                magwheels.append(magwheel_data)
            
            request.session['unsaved_magwheels'] = magwheels
            request.session.modified = True
            return redirect('inventory:v2-create-transaction')
    
    else:
        if index is not None and 0 <= int(index) < len(magwheels):
            form = MagwheelForm(initial = magwheels[int(index)])
        else:
            form = MagwheelForm()

    context = {
        'form': form,
        'edit_mode': index is not None,
        'magwheelTransaction': magwheels[int(index)] if index is not None and 0 <= int(index) < len(magwheels) else {},
        'brand_choices': MAGWHEELS_BRAND_CHOICES,
        'brand_size_choices': json.dumps(MAGWHEELS_BRAND_SIZE_CHOICES),
        'model_pcd1_choices': json.dumps(MAGWHEELS_MODEL_PCD1_CHOICES),
        'model_pcd2_choices': json.dumps(MAGWHEELS_MODEL_PCD2_CHOICES),
        'model_offset_choices': json.dumps(MAGWHEELS_MODEL_OFFSET_CHOICES),
        'model_bore_choices': json.dumps(MAGWHEELS_MODEL_BORE_CHOICES),
        'model_color_choices': json.dumps(MAGWHEELS_MODEL_COLOR_CHOICES),
    }

    return render(request, 'inventory/v2_addMagwheel.html', context)

def v2_deleteMagwheel(request, index):
    magwheels = request.session.get('unsaved_magwheels', [])
    if 0 <= int(index) < len(magwheels):
        del magwheels[int(index)]
        request.session['unsaved_magwheels'] = magwheels
        request.session.modified = True
    return redirect('inventory:v2-create-transaction')

def v2_viewAllTires(request):
    tires = Tire_test.objects.select_related('transaction').order_by('-transaction__date', 'id')

    context = {
        'tires': tires,
    }
    return render(request,'inventory/v2_allTires.html', context)

def v2_editSavedTire(request, pk):
    tire = get_object_or_404(Tire_test, pk = pk)

    TIRES_BRAND_CHOICES = get_tires_brand_choices()
    TIRES_BRAND_SIZE_CHOICES = get_tires_brand_size_choices()
    TIRES_MODEL_SPECS_CHOICES = get_tires_model_specs_choices()
    TIRES_SPECS_LOAD_CHOICES = get_tires_specs_load_choices()
    TIRES_SPECS_REMARKS_CHOICES = get_tires_specs_remarks_choices()

    if request.method == "POST":
        form = SavedTireForm(request.POST, instance = tire)
        if form.is_valid():
            form.save()
            return redirect('inventory:v2-view-all-tires')
    else:
        form = SavedTireForm(instance = tire)

    context = {
        'form': form,
        'tireTransaction': tire,
        'brand_choices': TIRES_BRAND_CHOICES,
        'brand_size_choices': json.dumps(TIRES_BRAND_SIZE_CHOICES),
        'model_specs_choices': json.dumps(TIRES_MODEL_SPECS_CHOICES),
        'specs_load_choices': json.dumps(TIRES_SPECS_LOAD_CHOICES),
        'specs_remarks_choices': json.dumps(TIRES_SPECS_REMARKS_CHOICES),
    }

    return render(request, 'inventory/v2_editSavedTire.html', context)

def v2_deleteSavedTire(request, pk):
    tire = get_object_or_404(Tire_test, pk = pk)
    tire.delete()
    return redirect('inventory:v2-view-all-tires')

def v2_editSavedMagwheel(request, pk):
    magwheel = get_object_or_404(Magwheel_test, pk = pk)

    # Choices creation
    MAGWHEELS_BRAND_CHOICES = get_magwheels_brand_choices()
    MAGWHEELS_BRAND_SIZE_CHOICES = get_magwheels_brand_size_choices()
    MAGWHEELS_MODEL_PCD1_CHOICES = get_magwheels_model_pcd1_choices()
    MAGWHEELS_MODEL_PCD2_CHOICES = get_magwheels_model_pcd2_choices()
    MAGWHEELS_MODEL_OFFSET_CHOICES = get_magwheels_model_offset_choices()
    MAGWHEELS_MODEL_BORE_CHOICES = get_magwheels_model_bore_choices()
    MAGWHEELS_MODEL_COLOR_CHOICES = get_magwheels_model_color_choices()

    if request.method == "POST":
        form = SavedMagwheelForm(request.POST, instance = magwheel)
        if form.is_valid():
            form.save()
            return redirect('inventory:v2-view-all-magwheels')
    else:
        form = SavedMagwheelForm(instance = magwheel)
    
    context = {
        'form': form,
        'magwheelTransaction': magwheel,
        'brand_choices': MAGWHEELS_BRAND_CHOICES,
        'brand_size_choices': json.dumps(MAGWHEELS_BRAND_SIZE_CHOICES),
        'model_pcd1_choices': json.dumps(MAGWHEELS_MODEL_PCD1_CHOICES),
        'model_pcd2_choices': json.dumps(MAGWHEELS_MODEL_PCD2_CHOICES),
        'model_offset_choices': json.dumps(MAGWHEELS_MODEL_OFFSET_CHOICES),
        'model_bore_choices': json.dumps(MAGWHEELS_MODEL_BORE_CHOICES),
        'model_color_choices': json.dumps(MAGWHEELS_MODEL_COLOR_CHOICES),
    }

    return render(request, 'inventory/v2_editSavedMagwheel.html', context)

def v2_deleteSavedMagwheel(request, pk):
    magwheel = get_object_or_404(Magwheel_test, pk = pk)
    magwheel.delete()
    return redirect('inventory:v2-view-all-magwheels')

def v2_viewAllMagwheels(request):
    magwheels = Magwheel_test.objects.select_related('transaction').order_by('-transaction__date', 'id')

    context = {
        'magwheels': magwheels,
    }
    return render(request, 'inventory/v2_allMagwheels.html', context)

def v2_viewSummaryTires(request):
    tires = (
        Tire_test.objects
        .select_related('transaction')  # Efficiently pull related Transaction data
        .annotate(
            inventory_effect=Case(
                When(transaction__transaction_type='IN', then=F('qty')),
                When(transaction__transaction_type='OUT', then=-F('qty')),
                default=0,
                output_field=IntegerField()
            ),
            transaction_date=F('transaction__date'),
            transaction_location=F('transaction__location'),
        )
    )

    grouped_inventory = (
        tires
        .values(
            'transaction_location',  # location from transaction
            'brand', 'size', 'model', 'specs',
            'supplier', 'load', 'remarks'
        )
        .annotate(
            total_inventory=Sum('inventory_effect'),
            last_transaction_date=Max('transaction_date')
        )
        .order_by(
            'transaction_location', 'brand', 'size', 'model', 'specs'
        )
    )

    context = {
        'grouped_inventory': grouped_inventory,
    }

    return render(request, 'inventory/v2_viewSummaryTires.html', context)

def v2_viewSummaryMagwheels(request):
    magwheels = (
        Magwheel_test.objects
        .select_related('transaction')  # Efficiently pull related Transaction data
        .annotate(
            inventory_effect=Case(
                When(transaction__transaction_type='IN', then=F('qty')),
                When(transaction__transaction_type='OUT', then=-F('qty')),
                default=0,
                output_field=IntegerField()
            ),
            transaction_date=F('transaction__date'),
            transaction_location=F('transaction__location'),
        )
    )

    grouped_inventory = (
        magwheels
        .values(
            'transaction_location', 
            'brand', 'size', 'model', 'pcd1', 'pcd2', 'offset', 'bore', 'color', 'supplier'
        )
        .annotate(
            total_inventory=Sum('inventory_effect'),
            last_transaction_date=Max('transaction_date')
        )
        .order_by('transaction_location', 'brand', 'size', 'model', 'pcd1', 'offset', 'color')  # Optional for consistent grouping order
    )

    context = {
        'grouped_inventory': grouped_inventory,
    }

    return render(request, 'inventory/v2_viewSummaryMagwheels.html', context)

def v2_queryTires(request):
    # Start with all tires and select related transactions
    tires = Tire_test.objects.select_related('transaction').all()

    # Define default filter values
    filters = {
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'location': request.GET.get('location', ''),
        'transaction_type': request.GET.get('transaction_type', ''),
        'brand': request.GET.get('brand', ''),
        'size': request.GET.get('size', ''),
        'model': request.GET.get('model', ''),
        'specs': request.GET.get('specs', ''),
        'supplier': request.GET.get('supplier', ''),
        'notes': request.GET.get('notes', ''),
    }

    # Filter by transaction date
    if filters['start_date'] and filters['end_date']:
        tires = tires.filter(transaction__date__range=[filters['start_date'], filters['end_date']])
    elif filters['start_date']:
        tires = tires.filter(transaction__date__gte=filters['start_date'])
    elif filters['end_date']:
        tires = tires.filter(transaction__date__lte=filters['end_date'])

    # Apply filters using transaction and tire fields
    if filters['location']:
        tires = tires.filter(transaction__location__icontains=filters['location'])
    if filters['transaction_type']:
        tires = tires.filter(transaction__transaction_type__icontains=filters['transaction_type'])
    if filters['brand']:
        tires = tires.filter(brand__icontains=filters['brand'])
    if filters['size']:
        tires = tires.filter(size__icontains=filters['size'])
    if filters['model']:
        tires = tires.filter(model__icontains=filters['model'])
    if filters['specs']:
        tires = tires.filter(specs__icontains=filters['specs'])
    if filters['supplier']:
        tires = tires.filter(supplier__icontains=filters['supplier'])

    # Sort by transaction date (earliest to latest)
    tires = tires.order_by('transaction__date')

    # Initialize cumulative inventory
    cumulative_qty = {}
    tires_list = []

    for tire in tires:
        transaction = tire.transaction
        key = (
            transaction.location,
            tire.brand,
            tire.size,
            tire.model,
            tire.specs,
            tire.load,
            tire.supplier
        )

        # Determine quantity change based on transaction type
        if transaction.transaction_type == "IN":
            qty_change = tire.qty
        elif transaction.transaction_type == "OUT":
            qty_change = -tire.qty
        else:
            qty_change = 0

        # Initialize or update cumulative quantity
        cumulative_qty[key] = cumulative_qty.get(key, 0) + qty_change

        tires_list.append({
            'tire': tire,
            'transaction': transaction,
            'inventory_qty': cumulative_qty[key],
        })

    context = {
        'tires': tires_list[::-1],  # Reverse to show most recent first
        'filters': filters,
    }

    return render(request, 'inventory/v2_queryTires.html', context)

def v2_queryMagwheels(request):
    # Start with all magwheels and select related transactions
    magwheels = Magwheel_test.objects.select_related('transaction').all()

    # Define default filter values
    filters = {
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'location': request.GET.get('location', ''),
        'transaction_type': request.GET.get('transaction_type', ''),
        'brand': request.GET.get('brand', ''),
        'size': request.GET.get('size', ''),
        'model': request.GET.get('model', ''),
        'pcd1': request.GET.get('pcd1', ''),
        'pcd2': request.GET.get('pcd2', ''),
        'offset': request.GET.get('offset', ''),
        'color': request.GET.get('color', ''),
        'supplier': request.GET.get('supplier', ''),
        'notes': request.GET.get('notes', ''),
    }

    # Filter by transaction date range
    if filters['start_date'] and filters['end_date']:
        magwheels = magwheels.filter(transaction__date__range=[filters['start_date'], filters['end_date']])
    elif filters['start_date']:
        magwheels = magwheels.filter(transaction__date__gte=filters['start_date'])
    elif filters['end_date']:
        magwheels = magwheels.filter(transaction__date__lte=filters['end_date'])

    # Apply filters using related transaction or magwheel fields
    if filters['location']:
        magwheels = magwheels.filter(transaction__location__icontains=filters['location'])
    if filters['transaction_type']:
        magwheels = magwheels.filter(transaction__transaction_type__icontains=filters['transaction_type'])
    if filters['brand']:
        magwheels = magwheels.filter(brand__icontains=filters['brand'])
    if filters['size']:
        magwheels = magwheels.filter(size__icontains=filters['size'])
    if filters['model']:
        magwheels = magwheels.filter(model__icontains=filters['model'])
    if filters['pcd1']:
        magwheels = magwheels.filter(pcd1__icontains=filters['pcd1'])
    if filters['pcd2']:
        magwheels = magwheels.filter(pcd2__icontains=filters['pcd2'])
    if filters['offset']:
        magwheels = magwheels.filter(offset__icontains=filters['offset'])
    if filters['color']:
        magwheels = magwheels.filter(color__icontains=filters['color'])
    if filters['supplier']:
        magwheels = magwheels.filter(supplier__icontains=filters['supplier'])

    # Sort by transaction date (earliest first)
    magwheels = magwheels.order_by('transaction__date')

    # Initialize cumulative quantity tracking
    cumulative_qty = {}
    magwheels_list = []

    for magwheel in magwheels:
        transaction = magwheel.transaction

        # Create a key based on all grouping fields + location
        key = (
            transaction.location,
            magwheel.brand,
            magwheel.size,
            magwheel.model,
            magwheel.pcd1,
            magwheel.pcd2,
            magwheel.offset,
            magwheel.bore,
            magwheel.color
        )

        # Quantity change based on transaction type
        if transaction.transaction_type == "IN":
            qty_change = magwheel.qty
        elif transaction.transaction_type == "OUT":
            qty_change = -magwheel.qty
        else:
            qty_change = 0

        # Update cumulative quantity
        cumulative_qty[key] = cumulative_qty.get(key, 0) + qty_change

        # Append enriched data for display
        magwheels_list.append({
            'magwheel': magwheel,
            'transaction': transaction,
            'inventory_qty': cumulative_qty[key]
        })

    # Prepare context
    context = {
        'magwheels': magwheels_list[::-1],  # Show latest first
        'filters': filters,
    }

    return render(request, 'inventory/v2_queryMagwheels.html', context)

def v2_addPayment(request, index = None):
    payments = request.session.get('unsaved_payments', [])

    if request.method == "POST":
        form = TransactionPaymentForm(request.POST)
        if form.is_valid():
            payment_data = form.cleaned_data

            # Convert Decimal fields to float to avoid JSON serialization issues
            payment_data['amount'] = float(payment_data['amount'])

            if index is not None:
                payments[int(index)] = payment_data
            else:
                payments.append(payment_data)
            
            request.session['unsaved_payments'] = payments
            request.session.modified = True
            return redirect('inventory:v2-create-transaction')
        
    else:
        if index is not None and 0 <= int(index) < len(payments):
            form = TransactionPaymentForm(initial=payments[int(index)])
        else:
            form = TransactionPaymentForm()

    context = {
        'form': form,
        'edit_mode': index is not None,
        'payment': payments[int(index)] if index is not None and 0 <= int(index) < len(payments) else {},
    }

    return render(request, 'inventory/v2_addPayment.html', context)

def v2_deletePayment(request, index):
    payments = request.session.get('unsaved_payments', [])
    if 0 <= int(index) < len(payments):
        del payments[int(index)]
        request.session['unsaved_payments'] = payments
        request.session.modified = True
    return redirect('inventory:v2-create-transaction')

def v2_editSavedPayment(request, transaction_id, pk):
    payment = get_object_or_404(TransactionPayment, pk = pk, transaction_id = transaction_id)

    if request.method == 'POST':
        form = SavedPaymentForm(request.POST, instance = payment)
        if form.is_valid():
            new_status = form.cleaned_data.get('status')
            if payment.locked and not new_status:
                form.add_error('status', "Cannot uncheck a locked payment.")
            else:
                form.save()
                return redirect('inventory:v2-expand-transaction', pk=transaction_id)
    else:
        form = SavedPaymentForm(instance = payment)
    
    context = {
        'form':form,
        'payment': payment,
    }

    return render(request, 'inventory/v2_editSavedPayment.html', context)

def is_admin(user):
    return user.is_authenticated and user.is_staff # or user.is_superuser

@login_required
@user_passes_test(is_admin)
def v2_paymentReportView(request):

    # Step 1: Get filters from GET parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    locations = request.GET.getlist('locations')  # list of locations

    # Step 2: Base queryset
    payments = TransactionPayment.objects.select_related('transaction')

    # Step 3: Apply filters if present
    if start_date:
        payments = payments.filter(transaction__date__gte=start_date)
    if end_date:
        payments = payments.filter(transaction__date__lte=end_date)
    if locations:
        payments = payments.filter(transaction__location__in=locations)

    # Step 4: Prepare summary per location and payment method
    report = {
        loc: {
            method: {
                'Paid': 0,
                'Unpaid': 0,
                'Discount': 0,
                'Net': 0
            } for method in payment_method_choices
        } for loc in locations
    }

    bank_breakdown = {
        loc: {
            'Credit Card': {
                bank: {'Paid': 0, 'Unpaid': 0, 'Discount': 0, 'Net': 0}
                for bank in banks
            },
            'Debit Card': {
                bank: {'Paid': 0, 'Unpaid': 0, 'Discount': 0, 'Net': 0}
                for bank in banks
            },
        } for loc in locations
    }


    # Step 5: Populate actual payment data
    for payment in payments:
        loc = payment.transaction.location
        method = payment.method
        status = 'Paid' if payment.status else 'Unpaid'
        amount = float(payment.amount)

        if loc in report and method in report[loc]:
            report[loc][method][status] += amount

            # Only apply discounts to 'Paid'
            if status == 'Paid':
                discount = 0

                if method == 'Credit Card':
                    card_type = payment.credit_card_type or ''
                    for bank in banks:
                        if card_type.startswith(bank):
                            rate = cc_rate_dict.get(card_type, 0)
                            discount = amount * rate
                            break  # Only match one bank

                    # Update main report
                    report[loc][method]['Discount'] += discount
                    report[loc][method]['Net'] += (amount - discount)

                    # Always update original bank-level breakdown (with or without discount)
                    for bank in banks:
                        if card_type.startswith(bank):
                            bank_breakdown[loc]['Credit Card'][bank]['Paid'] += amount
                            bank_breakdown[loc]['Credit Card'][bank].setdefault('Discount', 0)
                            bank_breakdown[loc]['Credit Card'][bank].setdefault('Net', 0)
                            bank_breakdown[loc]['Credit Card'][bank]['Discount'] += discount
                            bank_breakdown[loc]['Credit Card'][bank]['Net'] += (amount - discount)
                            break

                elif method == 'Debit Card':
                    bank_name = payment.debit_card_bank or ''
                    for bank in banks:
                        if bank_name.startswith(bank):
                            rate = dc_rate_dict.get(bank_name, 0)
                            discount = amount * rate
                            break

                    report[loc][method]['Discount'] += discount
                    report[loc][method]['Net'] += (amount - discount)

                    for bank in banks:
                        if bank_name.startswith(bank):
                            bank_breakdown[loc]['Debit Card'][bank]['Paid'] += amount
                            bank_breakdown[loc]['Debit Card'][bank].setdefault('Discount', 0)
                            bank_breakdown[loc]['Debit Card'][bank].setdefault('Net', 0)
                            bank_breakdown[loc]['Debit Card'][bank]['Discount'] += discount
                            bank_breakdown[loc]['Debit Card'][bank]['Net'] += (amount - discount)
                            break

                elif method == 'BDO Checkout':
                    rate = 0.04
                    discount = amount * rate
                    report[loc][method]['Discount'] += discount
                    report[loc][method]['Net'] += (amount - discount)

                else:
                    # No discount, just count net
                    report[loc][method]['Net'] += amount

            else:
                # Unpaid handling (no discount)
                if method in ['Credit Card', 'Debit Card']:
                    # Preserve bank-level unpaid
                    if method == "Credit Card":
                        card_type = payment.credit_card_type or ''
                        for bank in banks:
                            if card_type.startswith(bank):
                                bank_breakdown[loc]["Credit Card"][bank]['Unpaid'] += amount
                                break
                    elif method == "Debit Card":
                        bank_name = payment.debit_card_bank or ''
                        for bank in banks:
                            if bank_name.startswith(bank):
                                bank_breakdown[loc]["Debit Card"][bank]['Unpaid'] += amount
                                break



    # Step 6: Compute location-level and overall totals
    location_totals = {}
    grand_totals = {'Paid': 0, 'Unpaid': 0, 'Total': 0}

    for loc, methods in report.items():
        paid_sum = sum(vals['Paid'] for vals in methods.values())
        unpaid_sum = sum(vals['Unpaid'] for vals in methods.values())
        discount_sum = sum(vals['Discount'] for vals in methods.values())
        net_sum = sum(vals['Net'] for vals in methods.values())
        total = paid_sum + unpaid_sum
        
        location_totals[loc] = {
            'Paid': paid_sum,
            'Unpaid': unpaid_sum,
            'Discount': discount_sum,
            'Net': net_sum,
            'Total': total
        }

        grand_totals['Paid'] += paid_sum
        grand_totals['Unpaid'] += unpaid_sum
        grand_totals['Total'] += total


    context = {
        'report': report,
        'bank_breakdown': bank_breakdown,
        'start_date': start_date,
        'end_date': end_date,
        'today': date.today().isoformat(),
        'selected_locations': locations,
        'all_locations': all_locations,
        'location_totals': location_totals,
        'grand_totals': grand_totals,
    }

    return render(request, 'inventory/v2_paymentReportView.html', context)

@login_required
@user_passes_test(is_admin)
def v2_paymentReportExport(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    locations = request.GET.getlist('locations')

    payments = TransactionPayment.objects.select_related('transaction')

    if start_date:
        payments = payments.filter(transaction__date__gte=start_date)
    if end_date:
        payments = payments.filter(transaction__date__lte=end_date)
    if locations:
        payments = payments.filter(transaction__location__in=locations)

    payment_methods = payment_method_choices
    banks = ['BDO', 'BPI', 'MetroBank']  # Can be expanded

    # Initialize report and bank breakdown with full fields
    report = {
        loc: {
            method: {'Paid': 0, 'Unpaid': 0, 'Discount': 0, 'Net': 0}
            for method in payment_methods
        } for loc in locations
    }

    bank_breakdown = {
        loc: {
            'Credit Card': {
                bank: {'Paid': 0, 'Unpaid': 0, 'Discount': 0, 'Net': 0} for bank in banks
            },
            'Debit Card': {
                bank: {'Paid': 0, 'Unpaid': 0, 'Discount': 0, 'Net': 0} for bank in banks
            }
        } for loc in locations
    }

    for payment in payments:
        loc = payment.transaction.location
        method = payment.method
        status = 'Paid' if payment.status else 'Unpaid'
        amount = float(payment.amount)

        if loc not in report or method not in report[loc]:
            continue

        report[loc][method][status] += amount

        if status == 'Paid':
            discount = 0

            if method == 'Credit Card':
                card_type = payment.credit_card_type or ''
                for bank in banks:
                    if card_type.startswith(bank):
                        rate = cc_rate_dict.get(card_type, 0)
                        discount = amount * rate
                        break
                report[loc][method]['Discount'] += discount
                report[loc][method]['Net'] += (amount - discount)

                for bank in banks:
                    if card_type.startswith(bank):
                        bank_breakdown[loc]['Credit Card'][bank]['Paid'] += amount
                        bank_breakdown[loc]['Credit Card'][bank]['Discount'] += discount
                        bank_breakdown[loc]['Credit Card'][bank]['Net'] += (amount - discount)
                        break

            elif method == 'Debit Card':
                bank_name = payment.debit_card_bank or ''
                for bank in banks:
                    if bank_name.startswith(bank):
                        rate = dc_rate_dict.get(bank_name, 0)
                        discount = amount * rate
                        break
                report[loc][method]['Discount'] += discount
                report[loc][method]['Net'] += (amount - discount)

                for bank in banks:
                    if bank_name.startswith(bank):
                        bank_breakdown[loc]['Debit Card'][bank]['Paid'] += amount
                        bank_breakdown[loc]['Debit Card'][bank]['Discount'] += discount
                        bank_breakdown[loc]['Debit Card'][bank]['Net'] += (amount - discount)
                        break

            elif method == 'BDO Checkout':
                rate = 0.04
                discount = amount * rate
                report[loc][method]['Discount'] += discount
                report[loc][method]['Net'] += (amount - discount)
            else:
                # No discount logic
                report[loc][method]['Net'] += amount
        else:
            if method == 'Credit Card':
                card_type = payment.credit_card_type or ''
                for bank in banks:
                    if card_type.startswith(bank):
                        bank_breakdown[loc]['Credit Card'][bank]['Unpaid'] += amount
                        break
            elif method == 'Debit Card':
                bank_name = payment.debit_card_bank or ''
                for bank in banks:
                    if bank_name.startswith(bank):
                        bank_breakdown[loc]['Debit Card'][bank]['Unpaid'] += amount
                        break

    # Compute totals
    location_totals = {
        loc: {
            'Paid': sum(report[loc][m]['Paid'] for m in payment_methods),
            'Unpaid': sum(report[loc][m]['Unpaid'] for m in payment_methods),
        }
        for loc in locations
    }
    for val in location_totals.values():
        val['Total'] = val['Paid'] + val['Unpaid']

    # Create Excel file
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for loc in locations:
        ws = wb.create_sheet(title=loc)

        ws.append([f"Payment Report for {loc}"])
        ws.append([f"From {start_date or 'N/A'} to {end_date or 'N/A'}"])
        ws.append([])

        headers = ["Payment Method", "Paid Amount", "Unpaid Amount", "Discount Amount", "Net Amount"]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_num).font = Font(bold=True)

        for method in payment_methods:
            data = report[loc][method]
            ws.append([method, data['Paid'], data['Unpaid'], data['Discount'], data['Net']])

            if method in ['Credit Card', 'Debit Card']:
                for bank in banks:
                    bank_data = bank_breakdown[loc][method][bank]
                    ws.append([
                        f"  → {bank}",
                        bank_data['Paid'],
                        bank_data['Unpaid'],
                        bank_data['Discount'],
                        bank_data['Net']
                    ])
                    for col in range(1, 6):
                        ws.cell(row=ws.max_row, column=col).font = Font(size=9)

        # Totals
        ws.append([])
        ws.append(["Total Paid", location_totals[loc]['Paid']])
        ws.append(["Total Unpaid", location_totals[loc]['Unpaid']])
        ws.append(["Overall Total", location_totals[loc]['Total']])
        for i in range(3):
            row = ws.max_row - 2 + i
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # Return Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"payment_report_{date.today().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
